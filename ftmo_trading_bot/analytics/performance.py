"""
===============================================================================
FTMO Trading Bot — Performance Analyzer (ระบบวิเคราะห์ผลงาน)
===============================================================================
คำนวณสถิติผลงานการเทรดแบบมืออาชีพ

สถิติที่คำนวณ:
1. พื้นฐาน: Win Rate, Total P/L, Average Win/Loss
2. ขั้นสูง: Profit Factor, Expectancy (R-Multiple)
3. ความเสี่ยง: Sharpe Ratio, Sortino Ratio, Max Drawdown
4. FTMO: Target Progress, DD ต่อเทียบกับ Limit

⚠️ คำนวณจากข้อมูลจริงทั้งหมด — ไม่ใช้ค่าประมาณ
===============================================================================
"""

import math
from datetime import datetime, date
from typing import Dict, List, Optional
from dataclasses import dataclass


@dataclass
class TradeResult:
    """ผลลัพธ์เทรดสำหรับคำนวณสถิติ"""
    ticket: int
    symbol: str
    trade_type: str         # "BUY" หรือ "SELL"
    entry_price: float
    close_price: float
    lot_size: float
    risk_amount: float      # จำนวนเงินที่เสี่ยง (USD)
    profit: float           # กำไร/ขาดทุน (USD)
    rr_achieved: float      # RR ที่ได้จริง (P/L ÷ Risk)
    open_time: datetime
    close_time: datetime
    confluence_score: float


class PerformanceAnalyzer:
    """
    ระบบวิเคราะห์ผลงานการเทรด

    ระดับสถิติ:
    1. Basic: Win/Loss, Win Rate, Average Win/Loss, Largest Win/Loss
    2. Advanced: Profit Factor, Expectancy, R-Multiple
    3. Risk: Sharpe Ratio, Sortino Ratio, Max Drawdown, Calmar Ratio
    4. Consistency: % Winning Days, Best/Worst Day, Streak
    5. FTMO-Specific: Target Progress, DD vs Limit
    """

    # จำนวนวันเทรดต่อปี (ใช้คำนวณ Annualized)
    TRADING_DAYS_PER_YEAR = 252

    def __init__(self, initial_balance: float = 100000.0):
        """
        เริ่มต้น Performance Analyzer

        Args:
            initial_balance: Balance เริ่มต้น (สำหรับคำนวณ %)
        """
        self._initial_balance = initial_balance
        self._trades: List[TradeResult] = []
        self._daily_pnl: Dict[str, float] = {}  # {date_str: daily_pnl}
        self._equity_curve: List[float] = [initial_balance]

        print("📈 [Performance] เริ่มต้นระบบวิเคราะห์ผลงาน")

    # =========================================================================
    # 📥 เพิ่มข้อมูลเทรด
    # =========================================================================

    def add_trade(self, trade_data: Dict):
        """
        เพิ่มเทรดที่ปิดแล้วเข้าสู่ระบบวิเคราะห์

        Args:
            trade_data: Dict จาก ExecutedTrade.to_dict()
        """
        risk_amount = trade_data.get("risk_amount", 1)
        profit = trade_data.get("profit", 0)
        rr_achieved = profit / risk_amount if risk_amount > 0 else 0

        # แปลงเวลา
        open_time = trade_data.get("open_time", "")
        close_time = trade_data.get("close_time", "")
        if isinstance(open_time, str) and open_time:
            try:
                open_time = datetime.fromisoformat(open_time)
            except (ValueError, TypeError):
                open_time = datetime.now()
        elif not isinstance(open_time, datetime):
            open_time = datetime.now()

        if isinstance(close_time, str) and close_time:
            try:
                close_time = datetime.fromisoformat(close_time)
            except (ValueError, TypeError):
                close_time = datetime.now()
        elif not isinstance(close_time, datetime):
            close_time = datetime.now()

        result = TradeResult(
            ticket=trade_data.get("ticket", 0),
            symbol=trade_data.get("symbol", ""),
            trade_type=trade_data.get("type", ""),
            entry_price=trade_data.get("entry_price", 0),
            close_price=trade_data.get("close_price", 0),
            lot_size=trade_data.get("lot_size", 0),
            risk_amount=risk_amount,
            profit=profit,
            rr_achieved=round(rr_achieved, 2),
            open_time=open_time,
            close_time=close_time,
            confluence_score=trade_data.get("confluence", 0),
        )

        self._trades.append(result)

        # อัพเดท Daily P/L
        trade_date = close_time.strftime("%Y-%m-%d") if isinstance(close_time, datetime) else str(date.today())
        self._daily_pnl[trade_date] = self._daily_pnl.get(trade_date, 0) + profit

        # อัพเดท Equity Curve
        new_equity = self._equity_curve[-1] + profit
        self._equity_curve.append(new_equity)

    def add_trades_batch(self, trades: List[Dict]):
        """เพิ่มเทรดหลายรายการพร้อมกัน"""
        for trade in trades:
            self.add_trade(trade)

    def load_from_excel(self, excel_path: str = None) -> int:
        """
        โหลดเทรดที่ปิดแล้วจากไฟล์ Excel (ftmo_trades.xlsx) เข้าสู่ analyzer
        ใช้ตอน startup เพื่อให้ equity curve / DD history คงความต่อเนื่อง
        ข้ามแถวที่ยังไม่ปิด (Close Time ว่าง)

        Args:
            excel_path: path ไฟล์ Excel (ดีฟอลต์ = logs/ftmo_trades.xlsx)

        Returns:
            int: จำนวนเทรดที่โหลดสำเร็จ
        """
        try:
            import openpyxl
        except ImportError:
            print("⚠️ [Performance] openpyxl ไม่พร้อม — ข้าม load_from_excel")
            return 0

        import os
        if excel_path is None:
            excel_path = os.path.join(os.getcwd(), "logs", "ftmo_trades.xlsx")

        if not os.path.exists(excel_path):
            print(f"ℹ️  [Performance] ไม่พบไฟล์ {excel_path} — เริ่มต้นจากศูนย์")
            return 0

        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            if "Trades" not in wb.sheetnames:
                print(f"⚠️ [Performance] ไฟล์ {excel_path} ไม่มี sheet 'Trades'")
                return 0

            ws = wb["Trades"]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                return 0

            headers = [str(h) if h is not None else "" for h in rows[0]]
            idx = {name: i for i, name in enumerate(headers)}

            def _get(row, key, default=None):
                i = idx.get(key)
                if i is None or i >= len(row):
                    return default
                return row[i] if row[i] is not None else default

            loaded = 0
            for row in rows[1:]:
                close_time = _get(row, "Close Time")
                if close_time in (None, "", "-"):
                    continue  # ข้ามเทรดที่ยังไม่ปิด

                try:
                    profit = float(_get(row, "P/L ($)", 0) or 0)
                    risk_amt = float(_get(row, "Risk$", 0) or 0)
                    if risk_amt <= 0:
                        # fallback: คำนวณจาก Risk% * initial balance
                        risk_pct = float(_get(row, "Risk%", 0) or 0)
                        risk_amt = self._initial_balance * (risk_pct / 100.0) if risk_pct > 0 else 1.0

                    trade_dict = {
                        "ticket": int(_get(row, "Ticket", 0) or 0),
                        "symbol": str(_get(row, "Symbol", "") or ""),
                        "type": str(_get(row, "Type", "") or ""),
                        "entry_price": float(_get(row, "Entry", 0) or 0),
                        "close_price": float(_get(row, "Close Price", 0) or 0),
                        "lot_size": float(_get(row, "Lot", 0) or 0),
                        "risk_amount": risk_amt,
                        "profit": profit,
                        "open_time": _get(row, "Open Time", ""),
                        "close_time": close_time,
                        "confluence": float(_get(row, "Confluence", 0) or 0),
                    }
                    self.add_trade(trade_dict)
                    loaded += 1
                except (ValueError, TypeError) as e:
                    print(f"⚠️ [Performance] ข้ามแถวเสีย: {e}")
                    continue

            wb.close()
            print(f"📥 [Performance] โหลด {loaded} เทรดจาก {excel_path}")
            return loaded

        except Exception as e:
            print(f"⚠️ [Performance] load_from_excel ล้มเหลว: {e}")
            return 0

    # =========================================================================
    # 📊 สถิติพื้นฐาน (Basic Stats)
    # =========================================================================

    def get_basic_stats(self) -> Dict:
        """
        สถิติพื้นฐาน: Win/Loss, Win Rate, Average Win/Loss

        Returns:
            Dict: สถิติพื้นฐาน
        """
        if not self._trades:
            return self._empty_stats("basic")

        total = len(self._trades)
        wins = [t for t in self._trades if t.profit > 0]
        losses = [t for t in self._trades if t.profit <= 0]
        breakevens = [t for t in self._trades if t.profit == 0]

        total_profit = sum(t.profit for t in self._trades)
        gross_profit = sum(t.profit for t in wins)
        gross_loss = sum(t.profit for t in losses)

        avg_win = gross_profit / len(wins) if wins else 0
        avg_loss = abs(gross_loss / len(losses)) if losses else 0

        largest_win = max(t.profit for t in self._trades)
        largest_loss = min(t.profit for t in self._trades)

        return {
            "total_trades": total,
            "wins": len(wins),
            "losses": len(losses),
            "breakevens": len(breakevens),
            "win_rate": len(wins) / total * 100 if total > 0 else 0,
            "total_profit": round(total_profit, 2),
            "gross_profit": round(gross_profit, 2),
            "gross_loss": round(gross_loss, 2),
            "avg_win": round(avg_win, 2),
            "avg_loss": round(avg_loss, 2),
            "largest_win": round(largest_win, 2),
            "largest_loss": round(largest_loss, 2),
            "avg_trade": round(total_profit / total, 2) if total > 0 else 0,
        }

    # =========================================================================
    # 📊 สถิติขั้นสูง (Advanced Stats)
    # =========================================================================

    def get_advanced_stats(self) -> Dict:
        """
        สถิติขั้นสูง: Profit Factor, Expectancy, R-Multiple

        Returns:
            Dict: สถิติขั้นสูง
        """
        if not self._trades:
            return self._empty_stats("advanced")

        wins = [t for t in self._trades if t.profit > 0]
        losses = [t for t in self._trades if t.profit <= 0]

        gross_profit = sum(t.profit for t in wins)
        gross_loss = abs(sum(t.profit for t in losses))

        # Profit Factor = Gross Profit / Gross Loss
        profit_factor = gross_profit / gross_loss if gross_loss > 0 else float('inf')

        # Expectancy (ค่าเฉลี่ยเทรด)
        avg_win = gross_profit / len(wins) if wins else 0
        avg_loss = gross_loss / len(losses) if losses else 0
        win_rate = len(wins) / len(self._trades) if self._trades else 0

        expectancy = (win_rate * avg_win) - ((1 - win_rate) * avg_loss)

        # R-Multiple (กำไร/ขาดทุน เทียบกับ Risk)
        r_multiples = [t.rr_achieved for t in self._trades]
        avg_r = sum(r_multiples) / len(r_multiples) if r_multiples else 0

        # Win/Loss Ratio
        win_loss_ratio = avg_win / avg_loss if avg_loss > 0 else float('inf')

        return {
            "profit_factor": round(profit_factor, 2),
            "expectancy": round(expectancy, 2),
            "avg_r_multiple": round(avg_r, 2),
            "win_loss_ratio": round(win_loss_ratio, 2),
            "edge_ratio": round(expectancy / avg_loss * 100, 2) if avg_loss > 0 else 0,
        }

    # =========================================================================
    # 📊 สถิติความเสี่ยง (Risk Stats)
    # =========================================================================

    def get_risk_stats(self) -> Dict:
        """
        สถิติความเสี่ยง: Sharpe, Sortino, Max DD, Calmar

        Returns:
            Dict: สถิติความเสี่ยง
        """
        if len(self._trades) < 2:
            return self._empty_stats("risk")

        # ผลตอบแทนรายเทรด (%)
        returns = [t.profit / self._initial_balance for t in self._trades]
        avg_return = sum(returns) / len(returns)

        # Std Dev
        variance = sum((r - avg_return) ** 2 for r in returns) / len(returns)
        std_dev = math.sqrt(variance) if variance > 0 else 0.0001

        # Sharpe Ratio (Annualized) — สมมุติ Risk-Free Rate = 0
        trades_per_day = max(1, len(self._trades) / max(1, len(self._daily_pnl)))
        annualized_factor = math.sqrt(self.TRADING_DAYS_PER_YEAR * trades_per_day)
        sharpe = (avg_return / std_dev) * annualized_factor if std_dev > 0 else 0

        # Sortino Ratio (ใช้เฉพาะ Downside Deviation)
        downside_returns = [r for r in returns if r < 0]
        if downside_returns:
            downside_var = sum(r ** 2 for r in downside_returns) / len(downside_returns)
            downside_dev = math.sqrt(downside_var)
            sortino = (avg_return / downside_dev) * annualized_factor if downside_dev > 0 else 0
        else:
            sortino = float('inf')

        # Max Drawdown จาก Equity Curve
        max_dd, max_dd_pct = self._calculate_max_drawdown()

        # Calmar Ratio = Annualized Return / Max Drawdown
        total_return_pct = (self._equity_curve[-1] - self._initial_balance) / self._initial_balance
        calmar = abs(total_return_pct / max_dd_pct) if max_dd_pct != 0 else 0

        return {
            "sharpe_ratio": round(sharpe, 2),
            "sortino_ratio": round(sortino, 2),
            "max_drawdown_usd": round(max_dd, 2),
            "max_drawdown_pct": round(max_dd_pct * 100, 2),
            "calmar_ratio": round(calmar, 2),
            "volatility": round(std_dev * 100, 4),
            "current_equity": round(self._equity_curve[-1], 2),
        }

    def _calculate_max_drawdown(self):
        """คำนวณ Max Drawdown จาก Equity Curve"""
        if len(self._equity_curve) < 2:
            return 0.0, 0.0

        peak = self._equity_curve[0]
        max_dd = 0.0
        max_dd_pct = 0.0

        for equity in self._equity_curve:
            if equity > peak:
                peak = equity

            dd = peak - equity
            dd_pct = dd / peak if peak > 0 else 0

            if dd > max_dd:
                max_dd = dd
            if dd_pct > max_dd_pct:
                max_dd_pct = dd_pct

        return max_dd, max_dd_pct

    # =========================================================================
    # 📊 สถิติความสม่ำเสมอ (Consistency Stats)
    # =========================================================================

    def get_consistency_stats(self) -> Dict:
        """
        สถิติความสม่ำเสมอ: Winning Days %, Streak, Best/Worst Day

        Returns:
            Dict: สถิติความสม่ำเสมอ
        """
        if not self._daily_pnl:
            return self._empty_stats("consistency")

        daily_values = list(self._daily_pnl.values())
        winning_days = sum(1 for v in daily_values if v > 0)
        losing_days = sum(1 for v in daily_values if v <= 0)
        total_days = len(daily_values)

        best_day = max(daily_values)
        worst_day = min(daily_values)
        avg_daily_pnl = sum(daily_values) / total_days

        # คำนวณ Win/Loss Streak
        current_streak = 0
        max_win_streak = 0
        max_loss_streak = 0
        temp_streak = 0

        for pnl in daily_values:
            if pnl > 0:
                if temp_streak > 0:
                    temp_streak += 1
                else:
                    temp_streak = 1
                max_win_streak = max(max_win_streak, temp_streak)
            else:
                if temp_streak < 0:
                    temp_streak -= 1
                else:
                    temp_streak = -1
                max_loss_streak = max(max_loss_streak, abs(temp_streak))

        return {
            "total_days": total_days,
            "winning_days": winning_days,
            "losing_days": losing_days,
            "winning_days_pct": round(winning_days / total_days * 100, 1) if total_days > 0 else 0,
            "best_day": round(best_day, 2),
            "worst_day": round(worst_day, 2),
            "avg_daily_pnl": round(avg_daily_pnl, 2),
            "max_win_streak": max_win_streak,
            "max_loss_streak": max_loss_streak,
        }

    # =========================================================================
    # 📊 สถิติ FTMO (FTMO-Specific)
    # =========================================================================

    def get_ftmo_stats(self) -> Dict:
        """
        สถิติเฉพาะสำหรับ FTMO Challenge

        Returns:
            Dict: สถิติ FTMO
        """
        current_equity = self._equity_curve[-1]
        total_profit = current_equity - self._initial_balance
        target_pct = 0.10  # FTMO Challenge Target = 10%
        target_amount = self._initial_balance * target_pct

        progress = total_profit / target_amount * 100 if target_amount > 0 else 0

        _, max_dd_pct = self._calculate_max_drawdown()

        # Daily DD
        worst_daily = min(self._daily_pnl.values()) if self._daily_pnl else 0
        worst_daily_pct = worst_daily / self._initial_balance * 100

        return {
            "target_amount": round(target_amount, 2),
            "current_profit": round(total_profit, 2),
            "target_progress_pct": round(progress, 1),
            "max_dd_vs_limit": f"{max_dd_pct * 100:.2f}% / 10%",
            "worst_daily_vs_limit": f"{abs(worst_daily_pct):.2f}% / 5%",
            "days_traded": len(self._daily_pnl),
            "remaining_to_target": round(max(0, target_amount - total_profit), 2),
            "on_track": progress >= (len(self._daily_pnl) / 30 * 100) if self._daily_pnl else False,
        }

    # =========================================================================
    # 📊 สถิติรายคู่เงิน (Per-Symbol)
    # =========================================================================

    def get_per_symbol_stats(self) -> Dict[str, Dict]:
        """
        สถิติแยกตามคู่เงิน

        Returns:
            Dict: {symbol: {wins, losses, win_rate, total_pnl, ...}}
        """
        symbol_groups: Dict[str, List[TradeResult]] = {}
        for t in self._trades:
            if t.symbol not in symbol_groups:
                symbol_groups[t.symbol] = []
            symbol_groups[t.symbol].append(t)

        result = {}
        for symbol, trades in symbol_groups.items():
            wins = sum(1 for t in trades if t.profit > 0)
            total_pnl = sum(t.profit for t in trades)
            result[symbol] = {
                "trades": len(trades),
                "wins": wins,
                "losses": len(trades) - wins,
                "win_rate": round(wins / len(trades) * 100, 1),
                "total_pnl": round(total_pnl, 2),
                "avg_pnl": round(total_pnl / len(trades), 2),
            }

        return result

    # =========================================================================
    # 🏆 รายงานรวม (Full Report)
    # =========================================================================

    def get_full_report(self) -> Dict:
        """
        รวมทุกสถิติเป็น Dict เดียว — ส่งให้ Excel Logger

        Returns:
            Dict: สถิติทั้งหมดในรูปแบบ key-value
        """
        basic = self.get_basic_stats()
        advanced = self.get_advanced_stats()
        risk = self.get_risk_stats()
        consistency = self.get_consistency_stats()
        ftmo = self.get_ftmo_stats()

        report = {}

        # === Basic ===
        report["📊 Total Trades"] = basic.get("total_trades", 0)
        report["✅ Wins"] = basic.get("wins", 0)
        report["❌ Losses"] = basic.get("losses", 0)
        report["🏆 Win Rate %"] = f"{basic.get('win_rate', 0):.1f}%"
        report["💰 Total P/L"] = f"${basic.get('total_profit', 0):,.2f}"
        report["📈 Avg Win"] = f"${basic.get('avg_win', 0):,.2f}"
        report["📉 Avg Loss"] = f"${basic.get('avg_loss', 0):,.2f}"
        report["🏅 Largest Win"] = f"${basic.get('largest_win', 0):,.2f}"
        report["💀 Largest Loss"] = f"${basic.get('largest_loss', 0):,.2f}"
        report["─"] = "─" * 30

        # === Advanced ===
        report["⚖️ Profit Factor"] = advanced.get("profit_factor", 0)
        report["🎯 Expectancy $/trade"] = f"${advanced.get('expectancy', 0):,.2f}"
        report["📐 Avg R-Multiple"] = advanced.get("avg_r_multiple", 0)
        report["📊 Win/Loss Ratio"] = advanced.get("win_loss_ratio", 0)
        report["──"] = "─" * 30

        # === Risk ===
        report["📈 Sharpe Ratio"] = risk.get("sharpe_ratio", 0)
        report["📈 Sortino Ratio"] = risk.get("sortino_ratio", 0)
        report["📉 Max Drawdown $"] = f"${risk.get('max_drawdown_usd', 0):,.2f}"
        report["📉 Max Drawdown %"] = f"{risk.get('max_drawdown_pct', 0):.2f}%"
        report["📈 Calmar Ratio"] = risk.get("calmar_ratio", 0)
        report["───"] = "─" * 30

        # === Consistency ===
        report["📅 Days Traded"] = consistency.get("total_days", 0)
        report["📈 Winning Days %"] = f"{consistency.get('winning_days_pct', 0):.1f}%"
        report["🏅 Best Day"] = f"${consistency.get('best_day', 0):,.2f}"
        report["💀 Worst Day"] = f"${consistency.get('worst_day', 0):,.2f}"
        report["🔥 Max Win Streak"] = consistency.get("max_win_streak", 0)
        report["❄️ Max Loss Streak"] = consistency.get("max_loss_streak", 0)
        report["────"] = "─" * 30

        # === FTMO ===
        report["🎯 FTMO Target"] = f"${ftmo.get('target_amount', 0):,.2f}"
        report["📊 Progress"] = f"{ftmo.get('target_progress_pct', 0):.1f}%"
        report["💰 Remaining"] = f"${ftmo.get('remaining_to_target', 0):,.2f}"
        report["📉 Max DD vs Limit"] = ftmo.get("max_dd_vs_limit", "N/A")
        report["📉 Worst Daily vs Limit"] = ftmo.get("worst_daily_vs_limit", "N/A")

        return report

    def print_report(self):
        """แสดงรายงานบน Console"""
        report = self.get_full_report()

        print("\n" + "=" * 60)
        print("🏆 === FTMO Trading Bot — Performance Report ===")
        print("=" * 60)

        for key, value in report.items():
            if key.startswith("─"):
                print("  " + str(value))
            else:
                print(f"  {key:<28} {value}")

        print("=" * 60 + "\n")

    # =========================================================================
    # 🛠️ ฟังก์ชันช่วย
    # =========================================================================

    def _empty_stats(self, category: str) -> Dict:
        """คืนค่าสถิติว่างเปล่า"""
        defaults = {
            "basic": {"total_trades": 0, "wins": 0, "losses": 0, "win_rate": 0, "total_profit": 0},
            "advanced": {"profit_factor": 0, "expectancy": 0, "avg_r_multiple": 0},
            "risk": {"sharpe_ratio": 0, "sortino_ratio": 0, "max_drawdown_usd": 0, "max_drawdown_pct": 0},
            "consistency": {"total_days": 0, "winning_days": 0, "winning_days_pct": 0},
        }
        return defaults.get(category, {})

    @property
    def trade_count(self) -> int:
        """จำนวนเทรดทั้งหมด"""
        return len(self._trades)

    @property
    def equity_curve(self) -> List[float]:
        """Equity Curve (สำหรับ Plot)"""
        return self._equity_curve.copy()

    def __repr__(self) -> str:
        basic = self.get_basic_stats()
        return (
            f"PerformanceAnalyzer(trades={basic['total_trades']}, "
            f"WR={basic['win_rate']:.1f}%, P/L=${basic['total_profit']:,.2f})"
        )
