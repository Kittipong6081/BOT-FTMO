"""
===============================================================================
FTMO Trading Bot — Trade Logger (ระบบบันทึกเทรดอัตโนมัติ)
===============================================================================
บันทึกทุกเทรดลงไฟล์ Excel (.xlsx) อัตโนมัติ

ความสามารถ:
1. บันทึกเทรดที่เปิด/ปิด ลง Sheet "Trades"
2. บันทึก Daily Summary ลง Sheet "Daily"
3. สร้างไฟล์ใหม่ทุกเดือน (เพื่อไม่ให้ไฟล์ใหญ่เกินไป)
4. Auto-format: สี, ความกว้างคอลัมน์, ตัวเลข

⚠️ ใช้ openpyxl — ไม่ต้องติดตั้ง Excel
===============================================================================
"""

import os
from datetime import date
from typing import Optional, Dict, List
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ [คำเตือน] ไม่พบ openpyxl — ระบบ Logger จะบันทึกเฉพาะ Console")

from config.settings import bot_config


class TradeLogger:
    """
    ระบบบันทึกเทรดลง Excel อัตโนมัติ

    โครงสร้างไฟล์ .xlsx:
    ├── Sheet "Trades"   — รายการเทรดทั้งหมด (1 แถว = 1 เทรด)
    ├── Sheet "Daily"    — สรุปรายวัน (1 แถว = 1 วัน)
    └── Sheet "Stats"    — สถิติรวม (Win Rate, Sharpe, etc.)

    คอลัมน์ใน Trades:
    Ticket | Symbol | Type | Entry | SL | TP | Lot | Risk% | RR |
    Confluence | Open Time | Close Price | Close Time | P/L | Reason | Close Reason
    """

    # === คอลัมน์สำหรับ Sheet "Trades" (Schema v3 — v6.9 enhanced) ===
    # v3 เพิ่ม fields สำหรับวิเคราะห์ live behavior of E1+E2 deployment:
    # ML score (cal/raw), agent decision, confluence breakdown, trade mgmt state,
    # bid/ask, market context (ADX H1/H4, MTF/D1 bias), account state
    # v8.0.6 (2026-05-07): SMC-specific columns removed
    # ❌ removed: HTF Bias, HTF/MTF/OB/FVG/Sweep pts, MTF Bias, D1 Bias (8 cols)
    # 66 → 58 cols. Schema migration auto-archives old xlsx.
    TRADE_HEADERS = [
        # --- core (cols 1-19) ---
        "Ticket", "Symbol", "Type", "Entry", "SL", "TP", "Lot",
        "Risk%", "Risk$", "RR", "Confluence", "ATR",
        "Open Time", "Close Price", "Close Time",
        "P/L ($)", "P/L (%)", "Close Reason", "Signal Reasons",
        # --- Time + spread context (cols 20-24) ---
        "Session", "DayOfWeek", "HourOfDay",
        "Spread@Entry", "Slippage",
        # --- Market context — MR-relevant only (col 25) ---
        "Volatility Regime",
        # --- Risk + outcome (cols 26-31) ---
        "ConsecLoss Before", "DD@Entry %",
        "MAE", "MFE",
        "Time-in-Trade (s)", "Exit Path",
        # --- ML / Agent decision (cols 32-36) ---
        "ML Score (cal)", "ML Score (raw)",
        "Agent Action", "Agent Decision", "ML Threshold",
        # --- Trade mgmt state (cols 37-40) ---
        "BE Moved", "Partial Closed", "Trailing", "Final SL",
        # --- Live execution details (cols 41-45) ---
        "Bid@Entry", "Ask@Entry", "Spread (pips)",
        "Bid@Exit", "Ask@Exit",
        # --- Market context (cols 46-47) — ADX is MR's trend filter ---
        "ADX H1", "ADX H4",
        # --- Account state (cols 48-50) ---
        "Balance@Entry", "Balance@Close", "Equity Peak",
        # --- Overtrading detection (cols 51-54) ---
        "Trades Today @Open", "Trades 1h @Open",
        "Sec Since Last Open", "Sec Since Last Same-Sym",
        # --- Trade mgmt skipped (col 55) ---
        "Partial Skipped",
        # --- Retrain capability — full obs vector at decision (col 56) ---
        "Obs JSON",
        # --- v7 Chronos forecast @ entry (cols 57-58) ---
        "Chronos Align", "Chronos Unc",
    ]

    # === คอลัมน์สำหรับ Sheet "Signals" (per-scan log) ===
    # ทุกการ scan ของ SMC strategy + agent decision = 1 row
    # ใช้ตรวจสอบ signal frequency, reject distribution, agent SKIP behavior ใน live
    # v8.0.6 (2026-05-07): SMC-specific columns removed
    # ❌ removed: HTF Bias, MTF Bias, D1 Bias (3 cols). 23 → 20 cols.
    SIGNAL_HEADERS = [
        "Time", "Symbol", "Direction", "Result",
        "Confluence", "ATR", "RR Target",
        "ML Score (cal)", "ML Score (raw)",
        "Agent Action", "Agent Decision", "ML Threshold",
        "ADX H1",                                  # MR uses ADX as trend filter
        "Session", "Spread (pips)",
        "Reject/Skip Reasons",
        "Executor Reject",
        "Obs JSON",
        "Chronos Align", "Chronos Unc",
    ]

    # === คอลัมน์สำหรับ Sheet "Daily" ===
    DAILY_HEADERS = [
        "Date", "Trades", "Wins", "Losses", "Win Rate%",
        "Gross Profit", "Gross Loss", "Net P/L",
        "Max DD%", "Daily DD%", "Balance EOD"
    ]

    # v8.0.6: name-based column lookup (1-indexed). Auto-computed from headers.
    # Replaces hardcoded `column=N` calls — schema changes auto-propagate.
    _COL = {name: i + 1 for i, name in enumerate(TRADE_HEADERS)}
    _SCOL = {name: i + 1 for i, name in enumerate(SIGNAL_HEADERS)}

    def __init__(self, log_dir: str = None):
        """
        เริ่มต้น Trade Logger

        Args:
            log_dir: โฟลเดอร์สำหรับเก็บไฟล์ Log (ค่าเริ่มต้น: ./logs/)
        """
        self._log_dir = log_dir or os.path.join(os.getcwd(), "logs")
        self._current_file: Optional[str] = None
        self._trade_count = 0
        self._daily_trades: List[Dict] = []  # เก็บเทรดของวันนี้

        # สร้างโฟลเดอร์ถ้ายังไม่มี
        os.makedirs(self._log_dir, exist_ok=True)

        print(f"📝 [Trade Logger] เริ่มต้นระบบบันทึกเทรด")
        print(f"   📂 โฟลเดอร์: {self._log_dir}")

    # =========================================================================
    # 📝 บันทึกเทรดที่เปิด
    # =========================================================================

    def log_trade_opened(self, trade_data: Dict):
        """
        บันทึกเทรดที่เปิดใหม่

        Args:
            trade_data: ข้อมูลเทรดจาก ExecutedTrade.to_dict()
        """
        self._trade_count += 1
        print(f"📝 [Logger] บันทึกเทรดเปิด #{self._trade_count}: "
              f"{trade_data.get('symbol', '?')} {trade_data.get('type', '?')} "
              f"Lot={trade_data.get('lot_size', 0):.2f}")

        if not OPENPYXL_AVAILABLE:
            return

        try:
            wb, filepath = self._get_or_create_workbook()
            ws = wb["Trades"]

            # เพิ่มแถวใหม่ (Schema v3 — v6.9 E1+E2 enhanced)
            row = [
                trade_data.get("ticket", 0),
                trade_data.get("symbol", ""),
                trade_data.get("type", ""),
                trade_data.get("entry_price", 0),
                trade_data.get("sl_price", 0),
                trade_data.get("tp_price", 0),
                trade_data.get("lot_size", 0),
                trade_data.get("risk_pct", 0),
                trade_data.get("risk_amount", 0),
                trade_data.get("rr_ratio", 0),
                trade_data.get("confluence", 0),
                trade_data.get("atr", 0),
                trade_data.get("open_time", ""),
                "",   # close_price (ยังไม่ปิด)
                "",   # close_time
                "",   # P/L $
                "",   # P/L %
                "",   # close_reason
                trade_data.get("reasons", ""),
                # --- ML features v2 ---
                trade_data.get("session", ""),
                trade_data.get("day_of_week", ""),
                trade_data.get("hour_of_day", ""),
                trade_data.get("spread_at_entry", 0),
                trade_data.get("slippage", 0),
                # v8.0.6: removed htf_bias (SMC-only)
                trade_data.get("volatility_regime", ""),
                trade_data.get("consec_loss_before", 0),
                trade_data.get("dd_at_entry_pct", 0),
                "",   # MAE (อัพเดตตอนปิด)
                "",   # MFE
                "",   # time_in_trade
                "",   # exit_path
                # --- ML / Agent decision context ---
                trade_data.get("ml_score", 0.5),
                trade_data.get("ml_score_raw", 0.5),
                trade_data.get("agent_action_value", 0),
                trade_data.get("agent_decision", ""),
                trade_data.get("ml_threshold_used", 0),
                # v8.0.6: removed htf_score, mtf_score, ob_pts, fvg_pts, sweep_pts (SMC-only)
                "",   # be_moved (อัพเดตตอนปิด)
                "",   # partial_closed_flag
                "",   # trailing_active
                "",   # final_sl_at_close
                trade_data.get("bid_at_entry", 0),
                trade_data.get("ask_at_entry", 0),
                trade_data.get("spread_pips_actual", 0),
                "",   # bid_at_exit
                "",   # ask_at_exit
                trade_data.get("adx_h1", 0),
                trade_data.get("adx_h4", 0),
                # v8.0.6: removed mtf_bias, d1_bias (SMC-only)
                trade_data.get("balance_at_entry", 0),
                "",   # balance_at_close
                "",   # equity_peak_during_trade
                # --- Overtrading metrics ---
                trade_data.get("trades_today_at_open", 0),
                trade_data.get("trades_last_hour_at_open", 0),
                trade_data.get("secs_since_last_trade_open", 0),
                trade_data.get("secs_since_last_trade_same_symbol", 0),
                # v6.10: partial close skipped flag (lot too small for partial)
                bool(trade_data.get("partial_close_skipped", False)),
                # --- Obs vector (JSON) — for offline retrain ---
                str(trade_data.get("obs_27_json", ""))[:600],
                # v7: Chronos forecast features @ entry (mirrors obs[27,28])
                round(float(trade_data.get("chronos_align", 0.0) or 0.0), 4),
                round(float(trade_data.get("chronos_unc", 0.0) or 0.0), 4),
            ]
            ws.append(row)

            # ใส่สี BUY=เขียว, SELL=แดง
            last_row = ws.max_row
            trade_type = trade_data.get("type", "")
            if trade_type == "BUY":
                ws.cell(row=last_row, column=3).fill = PatternFill("solid", fgColor="C6EFCE")
            elif trade_type == "SELL":
                ws.cell(row=last_row, column=3).fill = PatternFill("solid", fgColor="FFC7CE")

            wb.save(filepath)

        except Exception as e:
            print(f"⚠️ [Logger] บันทึกเทรดเปิดล้มเหลว: {e}")

    # =========================================================================
    # 📝 บันทึกเทรดที่ปิด
    # =========================================================================

    def log_trade_closed(self, trade_data: Dict):
        """
        อัพเดทเทรดที่ปิดแล้ว — เพิ่มราคาปิด, เวลาปิด, P/L

        Args:
            trade_data: ข้อมูลเทรดที่ปิดแล้ว
        """
        profit = trade_data.get("profit", 0)
        pnl_emoji = "🟢" if profit >= 0 else "🔴"
        print(f"{pnl_emoji} [Logger] บันทึกเทรดปิด: "
              f"{trade_data.get('symbol', '?')} P/L=${profit:,.2f}")

        # เก็บไว้สำหรับ Daily Summary
        self._daily_trades.append(trade_data)

        if not OPENPYXL_AVAILABLE:
            return

        try:
            wb, filepath = self._get_or_create_workbook()
            ws = wb["Trades"]

            ticket = trade_data.get("ticket", 0)

            # ค้นหาแถวที่ Ticket ตรงกัน
            target_row = None
            for row_idx in range(2, ws.max_row + 1):
                cell_val = ws.cell(row=row_idx, column=1).value
                if cell_val == ticket:
                    target_row = row_idx
                    break

            if target_row:
                # v8.0.6: name-based column lookup via self._COL — schema changes
                # auto-propagate, no more hardcoded column numbers + off-by-one bugs.
                C = self._COL
                ws.cell(row=target_row, column=C["Close Price"], value=trade_data.get("close_price", 0))
                ws.cell(row=target_row, column=C["Close Time"], value=trade_data.get("close_time", ""))
                ws.cell(row=target_row, column=C["P/L ($)"], value=profit)

                # คำนวณ P/L %
                risk_amount = trade_data.get("risk_amount", 1)
                pnl_pct = (profit / risk_amount * 100) if risk_amount > 0 else 0
                ws.cell(row=target_row, column=C["P/L (%)"], value=round(pnl_pct, 2))
                ws.cell(row=target_row, column=C["Close Reason"], value=trade_data.get("close_reason", ""))

                # ใส่สี P/L
                pnl_fill = PatternFill("solid", fgColor="C6EFCE") if profit >= 0 else PatternFill("solid", fgColor="FFC7CE")
                ws.cell(row=target_row, column=C["P/L ($)"]).fill = pnl_fill

                # --- ML features: MAE, MFE, time_in_trade, exit_path ---
                ws.cell(row=target_row, column=C["MAE"], value=trade_data.get("mae", 0))
                ws.cell(row=target_row, column=C["MFE"], value=trade_data.get("mfe", 0))
                ws.cell(row=target_row, column=C["Time-in-Trade (s)"], value=trade_data.get("time_in_trade", 0))
                ws.cell(row=target_row, column=C["Exit Path"],
                        value=trade_data.get("exit_path", trade_data.get("close_reason", "")))

                # --- Trade mgmt + close-time fields ---
                ws.cell(row=target_row, column=C["BE Moved"], value=bool(trade_data.get("be_moved", False)))
                ws.cell(row=target_row, column=C["Partial Closed"], value=bool(trade_data.get("partial_closed_flag", False)))
                ws.cell(row=target_row, column=C["Trailing"], value=bool(trade_data.get("trailing_active", False)))
                ws.cell(row=target_row, column=C["Final SL"], value=trade_data.get("final_sl_at_close", 0))
                ws.cell(row=target_row, column=C["Bid@Exit"], value=trade_data.get("bid_at_exit", 0))
                ws.cell(row=target_row, column=C["Ask@Exit"], value=trade_data.get("ask_at_exit", 0))
                ws.cell(row=target_row, column=C["Balance@Close"], value=trade_data.get("balance_at_close", 0))
                ws.cell(row=target_row, column=C["Equity Peak"], value=trade_data.get("equity_peak_during_trade", 0))
            else:
                # ไม่เจอ Ticket — เพิ่มแถวใหม่พร้อมข้อมูลครบ
                row = [
                    ticket,
                    trade_data.get("symbol", ""),
                    trade_data.get("type", ""),
                    trade_data.get("entry_price", 0),
                    trade_data.get("sl_price", 0),
                    trade_data.get("tp_price", 0),
                    trade_data.get("lot_size", 0),
                    trade_data.get("risk_pct", 0),
                    trade_data.get("risk_amount", 0),
                    trade_data.get("rr_ratio", 0),
                    trade_data.get("confluence", 0),
                    trade_data.get("atr", 0),
                    trade_data.get("open_time", ""),
                    trade_data.get("close_price", 0),
                    trade_data.get("close_time", ""),
                    profit,
                    0,
                    trade_data.get("close_reason", ""),
                    trade_data.get("reasons", ""),
                ]
                ws.append(row)

            wb.save(filepath)

        except Exception as e:
            print(f"⚠️ [Logger] บันทึกเทรดปิดล้มเหลว: {e}")

    # =========================================================================
    # 🔍 บันทึก Signal Scan (per-scan log — รวม SKIP / NO_SIGNAL / REJECTED)
    # =========================================================================

    def log_signal_scan(self, scan_data: Dict):
        """
        บันทึก signal scan event (ทุกครั้งที่ SMC strategy run, ไม่ว่า outcome อะไร)

        เป้าหมาย: เก็บ live signal frequency, reject distribution, agent SKIP behavior
        ในรูปแบบที่ vis ภายหลังได้ → เทียบ backtest

        Args:
            scan_data: Dict with keys:
                time, symbol, direction, result, confluence, atr, rr_target,
                ml_score, ml_score_raw, agent_action_value, agent_decision,
                ml_threshold, adx_h1, htf_bias, mtf_bias, d1_bias,
                session, spread_pips, reasons (list or str)

            result values: NO_SIGNAL / REJECTED / AGENT_SKIP / AGENT_TAKE / AGENT_TAKE_FAIL / ML_FILTERED (v6.12)
        """
        if not OPENPYXL_AVAILABLE:
            return

        try:
            wb, filepath = self._get_or_create_workbook()

            if "Signals" not in wb.sheetnames:
                self._create_signal_sheet(wb)

            ws = wb["Signals"]

            reasons = scan_data.get("reasons", "")
            if isinstance(reasons, list):
                reasons = "; ".join(str(r) for r in reasons)

            row = [
                str(scan_data.get("time", "")),
                scan_data.get("symbol", ""),
                scan_data.get("direction", ""),
                scan_data.get("result", ""),
                scan_data.get("confluence", 0),
                scan_data.get("atr", 0),
                scan_data.get("rr_target", 0),
                scan_data.get("ml_score", 0),
                scan_data.get("ml_score_raw", 0),
                scan_data.get("agent_action_value", 0),
                scan_data.get("agent_decision", ""),
                scan_data.get("ml_threshold", 0),
                scan_data.get("adx_h1", 0),
                # v8.0.6: removed htf_bias, mtf_bias, d1_bias (SMC-only)
                scan_data.get("session", ""),
                scan_data.get("spread_pips", 0),
                reasons[:500],
                str(scan_data.get("executor_reject_reason", ""))[:80],
                str(scan_data.get("obs_27_json", ""))[:600],
                round(float(scan_data.get("chronos_align", 0.0) or 0.0), 4),
                round(float(scan_data.get("chronos_unc", 0.0) or 0.0), 4),
            ]
            ws.append(row)

            # Color row by result
            last_row = ws.max_row
            result = scan_data.get("result", "")
            color_map = {
                "AGENT_TAKE": "C6EFCE",       # green
                "AGENT_SKIP": "FFEB9C",       # yellow
                "REJECTED": "FFC7CE",         # red
                "AGENT_TAKE_FAIL": "FFC7CE",  # red
                "NO_SIGNAL": "F2F2F2",        # grey
                "ML_FILTERED": "DDEBF7",      # light-blue (v6.12 — pre-agent ML gate)
            }
            if result in color_map:
                ws.cell(row=last_row, column=4).fill = PatternFill(
                    "solid", fgColor=color_map[result]
                )

            wb.save(filepath)

        except Exception as e:
            # Don't break bot loop on logger failure — print only
            print(f"⚠️ [Logger] log_signal_scan ล้มเหลว: {e}")

    def _create_signal_sheet(self, wb):
        """Setup Signals sheet — headers + formatting"""
        ws = wb.create_sheet("Signals")

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="6F4E7C")  # purple
        header_align = Alignment(horizontal="center", vertical="center")

        ws.append(self.SIGNAL_HEADERS)
        for col_idx, _ in enumerate(self.SIGNAL_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Column widths (matches SIGNAL_HEADERS length 23 — v7)
        col_widths = [
            20, 10, 6, 18,        # Time, Symbol, Direction, Result
            10, 10, 10,           # Confluence, ATR, RR Target
            12, 12,               # ML Score (cal/raw)
            12, 14, 12,           # Agent Action, Decision, ML Threshold
            8, 9, 9, 8,           # ADX H1, HTF/MTF/D1 Bias
            10, 10,               # Session, Spread
            60,                   # Reject/Skip Reasons
            30,                   # Executor Reject (v6.10)
            60,                   # Obs27 JSON
            10, 10,               # v7: Chronos Align, Chronos Unc
        ]
        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        ws.freeze_panes = "A2"

    # =========================================================================
    # 📊 บันทึก Daily Summary
    # =========================================================================

    def log_daily_summary(
        self,
        balance: float,
        daily_dd_pct: float = 0.0,
        max_dd_pct: float = 0.0,
    ):
        """
        บันทึกสรุปรายวัน — เรียกตอนสิ้นวัน

        Args:
            balance: Balance สิ้นวัน
            daily_dd_pct: Daily Drawdown %
            max_dd_pct: Max Drawdown %
        """
        today = date.today().isoformat()
        trades = self._daily_trades
        total = len(trades)
        wins = sum(1 for t in trades if t.get("profit", 0) > 0)
        losses = total - wins
        win_rate = (wins / total * 100) if total > 0 else 0

        gross_profit = sum(t.get("profit", 0) for t in trades if t.get("profit", 0) > 0)
        gross_loss = sum(t.get("profit", 0) for t in trades if t.get("profit", 0) <= 0)
        net_pnl = gross_profit + gross_loss

        print(f"\n📊 [Logger] === สรุปรายวัน {today} ===")
        print(f"   📈 เทรดทั้งหมด: {total} (ชนะ {wins}, แพ้ {losses})")
        print(f"   🏆 Win Rate: {win_rate:.0f}%")
        print(f"   💰 Net P/L: ${net_pnl:,.2f}")
        print(f"   💎 Balance: ${balance:,.2f}")

        if not OPENPYXL_AVAILABLE:
            self._daily_trades.clear()
            return

        try:
            wb, filepath = self._get_or_create_workbook()

            if "Daily" not in wb.sheetnames:
                self._create_daily_sheet(wb)

            ws = wb["Daily"]
            row = [
                today, total, wins, losses, round(win_rate, 1),
                round(gross_profit, 2), round(gross_loss, 2), round(net_pnl, 2),
                round(max_dd_pct * 100, 2), round(daily_dd_pct * 100, 2),
                round(balance, 2)
            ]
            ws.append(row)

            # ใส่สี Net P/L
            last_row = ws.max_row
            pnl_fill = PatternFill("solid", fgColor="C6EFCE") if net_pnl >= 0 else PatternFill("solid", fgColor="FFC7CE")
            ws.cell(row=last_row, column=8).fill = pnl_fill

            wb.save(filepath)
            print(f"   📂 บันทึกลง: {filepath}")

        except Exception as e:
            print(f"⚠️ [Logger] บันทึก Daily Summary ล้มเหลว: {e}")

        # เคลียร์เทรดของวันนี้
        self._daily_trades.clear()

    # =========================================================================
    # 📂 จัดการไฟล์ Excel
    # =========================================================================

    def _get_or_create_workbook(self):
        """
        ดึง Workbook ปัจจุบัน หรือสร้างใหม่ถ้ายังไม่มี

        ใช้ไฟล์คงที่ชื่อ ftmo_trades.xlsx (single consolidated file)
        เพื่อให้ ML อ่านประวัติทั้งหมดได้ในไฟล์เดียว

        Returns:
            Tuple[Workbook, str]: (Workbook, filepath)
        """
        filename = "ftmo_trades.xlsx"
        filepath = os.path.join(self._log_dir, filename)

        if os.path.exists(filepath):
            # v8.0.6: schema migration — if existing xlsx has wrong col count
            # (legacy v7 had 66 trades cols / 23 signals cols, v8 has 58/20),
            # auto-archive to *.bak_pre_v8 and create fresh. Prevents off-by-one
            # writes when SMC cols were removed.
            try:
                _wb_check = openpyxl.load_workbook(filepath, read_only=True)
                _ok = True
                if "Trades" in _wb_check.sheetnames:
                    if _wb_check["Trades"].max_column not in (0, len(self.TRADE_HEADERS)):
                        _ok = False
                if "Signals" in _wb_check.sheetnames:
                    if _wb_check["Signals"].max_column not in (0, len(self.SIGNAL_HEADERS)):
                        _ok = False
                _wb_check.close()
                if not _ok:
                    import time as _t
                    bak_path = filepath.replace(".xlsx", f".bak_pre_v8_{int(_t.time())}.xlsx")
                    os.rename(filepath, bak_path)
                    print(f"📦 [Logger] Schema mismatch — archived old xlsx → {os.path.basename(bak_path)}")
                    print(f"   (v8.0.6 removed SMC cols: 66→58 Trades / 23→20 Signals)")
            except Exception as _e:
                print(f"⚠️ [Logger] schema check failed ({_e}) — proceeding")

        if os.path.exists(filepath):
            wb = openpyxl.load_workbook(filepath)
        else:
            wb = self._create_new_workbook()
            wb.save(filepath)
            print(f"📂 [Logger] สร้างไฟล์ใหม่: {filepath}")

        self._current_file = filepath
        return wb, filepath

    def _create_new_workbook(self):
        """สร้าง Workbook ใหม่พร้อม Headers และ Formatting"""
        wb = openpyxl.Workbook()

        # === Sheet "Trades" ===
        ws_trades = wb.active
        ws_trades.title = "Trades"
        self._setup_trade_sheet(ws_trades)

        # === Sheet "Daily" ===
        ws_daily = wb.create_sheet("Daily")
        self._create_daily_sheet(wb)

        # === Sheet "Stats" ===
        ws_stats = wb.create_sheet("Stats")
        self._setup_stats_sheet(ws_stats)

        return wb

    def _setup_trade_sheet(self, ws):
        """ตั้งค่า Sheet Trades — Headers + Formatting"""
        # Header Style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2F5496")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # เพิ่ม Headers
        ws.append(self.TRADE_HEADERS)
        for col_idx, header in enumerate(self.TRADE_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # กำหนดความกว้างคอลัมน์ (Schema v3 — 64 คอลัมน์ พร้อม Partial Skipped + Obs27 JSON)
        col_widths = [
            # core (1-19)
            12, 10, 6, 10, 10, 10, 7, 7, 10, 6, 10, 10, 20, 10, 20, 12, 8, 20, 40,
            # v2 cols: Session, DoW, Hour, Spread, Slippage, HTF, VolRegime,
            # ConsecLoss, DD%, MAE, MFE, TimeInTrade, ExitPath (20-32)
            10, 6, 6, 10, 10, 8, 10, 10, 10, 10, 10, 14, 14,
            # v3 cols (33-56): ML/Agent (5), Confluence breakdown (5), Trade mgmt (4),
            # Live exec (5), Market context (4), Account state (3)
            12, 12, 12, 10, 10,                # ML/Agent (5)
            8, 8, 8, 8, 8,                     # Confluence breakdown (5)
            10, 14, 10, 10,                    # Trade mgmt (4)
            10, 10, 10, 10, 10,                # Live exec (5)
            8, 8, 9, 8,                        # Market context (4)
            12, 12, 12,                        # Account state (3)
            14, 14, 18, 20,                    # Overtrading metrics (4)
            12,                                # Partial Skipped (v6.10)
            60,                                # Obs27 JSON (1)
        ]
        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        # Freeze Header
        ws.freeze_panes = "A2"

    def _create_daily_sheet(self, wb):
        """ตั้งค่า Sheet Daily"""
        if "Daily" not in wb.sheetnames:
            wb.create_sheet("Daily")

        ws = wb["Daily"]

        # ถ้ายังไม่มี Header
        if ws.max_row <= 1 and ws.cell(row=1, column=1).value is None:
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill("solid", fgColor="548235")

            ws.append(self.DAILY_HEADERS)
            for col_idx, header in enumerate(self.DAILY_HEADERS, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            col_widths = [12, 8, 8, 8, 10, 12, 12, 12, 8, 10, 14]
            for idx, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(idx)].width = width

            ws.freeze_panes = "A2"

    def _setup_stats_sheet(self, ws):
        """ตั้งค่า Sheet Stats — สถิติรวม"""
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="BF8F00")

        stats_labels = [
            "Metric", "Value"
        ]
        ws.append(stats_labels)
        for col_idx, header in enumerate(stats_labels, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        ws.freeze_panes = "A2"

    # =========================================================================
    # 📊 อัพเดท Sheet Stats
    # =========================================================================

    def update_stats_sheet(self, stats: Dict):
        """
        อัพเดท Sheet "Stats" ด้วยสถิติล่าสุด

        Args:
            stats: Dict จาก PerformanceAnalyzer.get_full_report()
        """
        if not OPENPYXL_AVAILABLE:
            return

        try:
            wb, filepath = self._get_or_create_workbook()

            if "Stats" not in wb.sheetnames:
                ws = wb.create_sheet("Stats")
                self._setup_stats_sheet(ws)
            else:
                ws = wb["Stats"]

            # เคลียร์ข้อมูลเดิม (เก็บ Header)
            for row_idx in range(ws.max_row, 1, -1):
                ws.delete_rows(row_idx)

            # เพิ่มข้อมูลใหม่
            value_font = Font(size=11)
            label_font = Font(bold=True, size=11)

            for key, value in stats.items():
                ws.append([key, value])
                last_row = ws.max_row
                ws.cell(row=last_row, column=1).font = label_font
                ws.cell(row=last_row, column=2).font = value_font
                ws.cell(row=last_row, column=2).alignment = Alignment(horizontal="right")

            wb.save(filepath)

        except Exception as e:
            print(f"⚠️ [Logger] อัพเดท Stats ล้มเหลว: {e}")

    # =========================================================================
    # 📊 ข้อมูลสรุป
    # =========================================================================

    @property
    def current_file(self) -> Optional[str]:
        """ไฟล์ Log ปัจจุบัน"""
        return self._current_file

    @property
    def trade_count(self) -> int:
        """จำนวนเทรดที่บันทึก"""
        return self._trade_count

    def get_daily_trades(self) -> List[Dict]:
        """เทรดของวันนี้"""
        return self._daily_trades.copy()

    def __repr__(self) -> str:
        return f"TradeLogger(trades={self._trade_count}, file={self._current_file})"
